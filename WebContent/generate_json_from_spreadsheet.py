#!/usr/bin/env python
import urllib,sys, json
from openpyxl import load_workbook

def print_json(data):    
    with open('data/dive_site.json', 'w') as outfile:
        json.dump(data, outfile)
    
def import_stations(sheet,features):
    print 'Import stations name:'

    for i in range(1,50):
        cell_key='A'+str(i)
        if ( sheet[cell_key].value  is None ):
            continue
        
        feature={}
        feature['type']='Feature'        
        feature['id']=sheet[cell_key].value
        
        properties={}
        properties['id']=sheet[cell_key].value
        properties['description']=sheet['B'+str(i)].value
        properties['how_many_dives']='1'

        geometry={}
        geometry['type']='Point'
        coordinates=[]
        coordinates.append(sheet['G'+str(i)].value)
        coordinates.append(sheet['H'+str(i)].value)
        geometry['coordinates']=coordinates
        
        feature['properties']=properties
        feature['geometry']=geometry
        print (cell_key+' > '+sheet[cell_key].value)
        features.append(feature)

#    for row in sheet.iter_rows():
#        for cell in row:
#            #print(str(cell.row)+':'+str(cell.column)+' '+str(cell.value))
#            print(cell.value)


def main(argv):
    #------
    #--- Config parameter
    #-
    URL="https://docs.google.com/spreadsheets/d/1Nj7GsQPx4UWiOHRN7FY9XpiCSBuTekw_7dBzVcpHP4I/export?format=xlsx"
    path_excel='pippo.xlsx'

    stations_title='Stazioni'

    features= []
    dive_site = {}
    dive_site['type'] = 'FeatureCollection'
#
    #- Download the file
    urllib.urlretrieve (URL, path_excel)

    
    # Manage the spreadsheet
    wb = load_workbook(path_excel)

    #print wb.get_sheet_names()
    for sheet in wb:
        if stations_title == sheet.title:
            import_stations(sheet,features)

    #
    dive_site['features'] = features

    # print json format file
    print_json(dive_site)

   
if __name__ == "__main__":
    main(sys.argv)
        

    
